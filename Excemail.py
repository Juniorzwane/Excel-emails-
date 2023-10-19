import os
import re
import openpyxl
import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD

def extract_emails_from_cell(cell_value):
    email_pattern = r'\S+@\S+'
    emails = re.findall(email_pattern, cell_value)
    return emails

def process_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        with open('emails.txt', 'w') as email_file:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    for cell_value in row:
                        if isinstance(cell_value, str):
                            emails = extract_emails_from_cell(cell_value)
                            if emails:
                                email_file.write('\n'.join(emails) + '\n')
        result_label.config(text="Emails extracted and saved to emails.txt")
    except Exception as e:
        result_label.config(text=f"Error: {str(e)}")

def on_drop(event):
    file_path = event.data
    if file_path.endswith('.xlsx'):
        process_excel_file(file_path)
    else:
        result_label.config(text="Please drop an Excel file.")

root = TkinterDnD.Tk()
root.title("Excel Email Extractor")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack()

instruction_label = tk.Label(frame, text="Drag and drop an Excel file here:")
instruction_label.pack()

result_label = tk.Label(frame, text="")
result_label.pack()

root.drop_target_register(DND_FILES)
root.dnd_bind('<<Drop>>', on_drop)

root.mainloop()
